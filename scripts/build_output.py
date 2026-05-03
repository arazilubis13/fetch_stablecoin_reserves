"""
Builds the unified stablecoin reserve output workbook.

Reads all processed CSVs from data/processed/ and writes:
  output/stablecoin_reserves.xlsx

Sheets:
  One per coin  -- standardized aggregate reserves
  Securities    -- CUSIP-level detail (USDC + BUSD + PYUSD)
  USDTB_BUIDL  -- BUIDL holdings by blockchain

Usage:
    python scripts/build_output.py
"""

import csv
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT      = Path(__file__).parent.parent
PROCESSED = ROOT / "data" / "processed"
OUT_DIR   = ROOT / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_FILE  = OUT_DIR / "stablecoin_reserves.xlsx"

# ── schema ────────────────────────────────────────────────────────────────────

AGGREGATE_COLS = [
    "report_date", "source_file", "auditor",
    "tokens_in_circulation", "total_reserves", "excess_reserves",
    "treasuries", "repo", "money_mkt_funds", "cash", "settlement_adj",
    "secured_loans", "bitcoin", "gold", "non_us_treasuries", "buidl",
]

SECURITIES_COLS = [
    "report_date", "token", "source_file", "counterparty", "asset_type",
    "cusip", "maturity_date", "par_value", "fair_value", "days_to_maturity",
]

BUIDL_COLS = [
    "report_date", "source_file", "blockchain", "fair_value",
]


# ── CSV helpers ───────────────────────────────────────────────────────────────

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def to_num(val):
    if val in (None, "", "None"):
        return None
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return None


def to_date(val):
    if val in (None, "", "None"):
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    for fmt in ("%Y-%m-%d", "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


def empty_row(token):
    return {c: None for c in AGGREGATE_COLS} | {"_token": token}


# ── coin loaders ──────────────────────────────────────────────────────────────

def load_standard(path, token):
    """Coins already in the standard schema (manual-extracted coins)."""
    rows = []
    for r in read_csv(path):
        row = empty_row(token)
        for col in AGGREGATE_COLS:
            raw = r.get(col)
            if col == "report_date":
                row[col] = to_date(raw)
            elif col in ("source_file", "auditor"):
                row[col] = raw or None
            else:
                row[col] = to_num(raw)
        rows.append(row)
    return rows


def load_usdc(path):
    rows = []
    for r in read_csv(path):
        row = empty_row("USDC")
        row["report_date"]          = to_date(r.get("report_date"))
        row["source_file"]          = r.get("file")
        row["auditor"]              = r.get("auditor")
        row["tokens_in_circulation"]= to_num(r.get("usdc_in_circulation"))
        row["total_reserves"]       = to_num(r.get("total_reserves")) or to_num(r.get("fair_value_reserves"))
        row["treasuries"]           = to_num(r.get("treasury_securities"))
        row["repo"]                 = to_num(r.get("repo"))
        fund_cash  = to_num(r.get("fund_cash"))  or 0
        bank_cash  = to_num(r.get("bank_cash"))  or 0
        row["cash"]                 = (fund_cash + bank_cash) or None
        fund_set   = to_num(r.get("fund_settlement")) or 0
        other_set  = to_num(r.get("other_settlement")) or 0
        row["settlement_adj"]       = (fund_set + other_set) or None
        rows.append(row)
    return rows


def load_usdt(path):
    rows = []
    for r in read_csv(path):
        row = empty_row("USDT")
        row["report_date"]          = to_date(r.get("report_date"))
        row["source_file"]          = r.get("file")
        row["auditor"]              = r.get("auditor")
        row["tokens_in_circulation"]= to_num(r.get("usdt_in_circulation"))
        row["total_reserves"]       = to_num(r.get("total_assets"))
        row["excess_reserves"]      = to_num(r.get("excess_reserves"))
        row["treasuries"]           = to_num(r.get("us_treasuries"))
        row["repo"]                 = to_num(r.get("reverse_repos"))
        row["money_mkt_funds"]      = to_num(r.get("money_mkt_funds"))
        row["secured_loans"]        = to_num(r.get("secured_loans"))
        row["bitcoin"]              = to_num(r.get("bitcoin"))
        row["gold"]                 = to_num(r.get("gold"))
        rows.append(row)
    return rows


def load_fdusd(path):
    rows = []
    for r in read_csv(path):
        row = empty_row("FDUSD")
        row["report_date"]          = to_date(r.get("report_date"))
        row["source_file"]          = r.get("file")
        row["auditor"]              = r.get("auditor")
        row["tokens_in_circulation"]= to_num(r.get("fdusd_in_circulation"))
        row["total_reserves"]       = to_num(r.get("total_reserves"))
        row["treasuries"]           = to_num(r.get("us_treasuries"))
        row["repo"]                 = to_num(r.get("overnight_repo"))
        row["cash"]                 = to_num(r.get("cash"))
        rows.append(row)
    return rows


def load_rlusd(path):
    rows = []
    for r in read_csv(path):
        row = empty_row("RLUSD")
        row["report_date"]          = to_date(r.get("report_date"))
        row["source_file"]          = r.get("file")
        row["auditor"]              = r.get("auditor")
        row["tokens_in_circulation"]= to_num(r.get("rlusd_outstanding"))
        row["total_reserves"]       = to_num(r.get("total_reserves"))
        row["treasuries"]           = to_num(r.get("tbills"))
        row["money_mkt_funds"]      = to_num(r.get("money_mkt_funds"))
        row["cash"]                 = to_num(r.get("cash"))
        rows.append(row)
    return rows


def load_busd():
    """Pivot BUSD long-format summary CSVs into wide aggregate rows."""
    ASSET_MAP = {
        "treasuries":                    "treasuries",
        "treasury bill":                 "treasuries",
        "cash":                          "cash",
        "cash deposits with private insurance": "cash",
        "cash deposits at insured deposit institutions": "cash",
        "repo":                          "repo",
        "reverse repo collateral":       "repo",
        "overnight repo":                "repo",
    }

    summary_files = sorted(PROCESSED.glob("busd/busd_reserves_summary*.csv"))
    by_date = defaultdict(lambda: empty_row("BUSD"))

    for path in summary_files:
        for r in read_csv(path):
            d = to_date(r.get("date"))
            if not d:
                continue
            row = by_date[d]
            row["report_date"] = d
            row["source_file"] = path.name

            asset = str(r.get("asset_type") or "").strip().lower()
            col   = ASSET_MAP.get(asset)
            if col:
                existing = row.get(col) or 0
                row[col] = existing + (to_num(r.get("total_amount")) or 0)

    rows = sorted(by_date.values(), key=lambda r: r["report_date"] or date.min)

    # Compute total_reserves as sum of non-None asset cols
    asset_cols = ["treasuries", "repo", "cash", "money_mkt_funds",
                  "secured_loans", "bitcoin", "gold"]
    for row in rows:
        total = sum(row.get(c) or 0 for c in asset_cols)
        row["total_reserves"] = total or None

    return rows


def load_usdtb(path):
    rows = []
    for r in read_csv(path):
        row = empty_row("USDTB")
        row["report_date"]          = to_date(r.get("report_date"))
        row["source_file"]          = r.get("file")
        row["auditor"]              = r.get("auditor")
        row["tokens_in_circulation"]= to_num(r.get("tokens_in_circulation"))
        row["total_reserves"]       = to_num(r.get("total_reserves"))
        row["cash"]                 = to_num(r.get("cash"))
        row["buidl"]                = to_num(r.get("buidl"))
        row["excess_reserves"]      = to_num(r.get("surplus_deficit"))
        rows.append(row)
    return rows


# ── securities loaders ────────────────────────────────────────────────────────

def load_usdc_securities():
    rows = []
    path = PROCESSED / "usdc" / "usdc_treasury_securities.csv"
    if not path.exists():
        return rows
    for r in read_csv(path):
        rd = to_date(r.get("report_date"))
        mat = to_date(r.get("maturity_date"))
        rows.append({
            "report_date":    rd,
            "token":          "USDC",
            "source_file":    r.get("file"),
            "counterparty":   None,
            "asset_type":     "Treasury",
            "cusip":          r.get("cusip"),
            "maturity_date":  mat,
            "par_value":      None,
            "fair_value":     to_num(r.get("fair_value")),
            "days_to_maturity": (mat - rd).days if mat and rd else None,
        })
    return rows


def load_busd_securities():
    rows = []
    for path in sorted(PROCESSED.glob("busd/busd_reserves_detailed*.csv")):
        for r in read_csv(path):
            if not r.get("cusip"):
                continue
            rd  = to_date(r.get("date"))
            mat = to_date(r.get("maturity_date"))
            rows.append({
                "report_date":    rd,
                "token":          "BUSD",
                "source_file":    path.name,
                "counterparty":   None,
                "asset_type":     r.get("asset_type"),
                "cusip":          r.get("cusip"),
                "maturity_date":  mat,
                "par_value":      None,
                "fair_value":     to_num(r.get("amount")),
                "days_to_maturity": (mat - rd).days if mat and rd else None,
            })
    return rows


def load_pyusd_securities():
    rows = []
    path = PROCESSED / "pyusd" / "pyusd_securities.csv"
    if not path.exists():
        return rows
    for r in read_csv(path):
        rd  = to_date(r.get("report_date"))
        mat = to_date(r.get("maturity_date"))
        rows.append({
            "report_date":    rd,
            "token":          "PYUSD",
            "source_file":    r.get("source_file"),
            "counterparty":   r.get("counterparty"),
            "asset_type":     r.get("asset_type"),
            "cusip":          r.get("cusip"),
            "maturity_date":  mat,
            "par_value":      to_num(r.get("par_value")),
            "fair_value":     to_num(r.get("fair_value")),
            "days_to_maturity": (mat - rd).days if mat and rd else None,
        })
    return rows


# ── coin registry ─────────────────────────────────────────────────────────────

def load_all_coins():
    """Returns dict of {sheet_name: [rows]} for aggregate sheets."""
    coins = {}

    def _std(token, filename):
        path = PROCESSED / token.lower() / filename
        if path.exists():
            coins[token] = load_standard(path, token)

    coins["USDC"]  = load_usdc(PROCESSED / "usdc"  / "usdc_reserves.csv")
    coins["USDT"]  = load_usdt(PROCESSED / "usdt"  / "usdt_reserves.csv")
    coins["FDUSD"] = load_fdusd(PROCESSED / "fdusd" / "fdusd_reserves.csv")
    coins["RLUSD"] = load_rlusd(PROCESSED / "rlusd" / "rlusd_reserves.csv")
    coins["BUSD"]  = load_busd()
    coins["USDTB"] = load_usdtb(PROCESSED / "usdtb" / "usdtb_reserves.csv")
    coins["PYUSD"] = load_standard(PROCESSED / "pyusd" / "pyusd_reserves.csv", "PYUSD")

    for token, fname in [
        ("EURC",  "eurc_reserves.csv"),
        ("EURS",  "eurs_reserves.csv"),
        ("GUSD",  "gusd_reserves.csv"),
        ("USD1",  "usd1_reserves.csv"),
        ("USDP",  "usdp_reserves.csv"),
        ("AUDD",  "audd_reserves.csv"),
        ("AUDM",  "audm_reserves.csv"),
        ("TGBP",  "tgbp_reserves.csv"),
        ("VGBP",  "vgbp_reserves.csv"),
        ("XSGD",  "xsgd_reserves.csv"),
    ]:
        _std(token, fname)

    return coins


# ── Excel writer ──────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
DATE_FMT     = "YYYY-MM-DD"
NUM_FMT      = "#,##0.00"
INT_FMT      = "#,##0"

COL_WIDTHS = {
    "report_date": 13, "source_file": 28, "auditor": 18,
    "tokens_in_circulation": 22, "total_reserves": 18, "excess_reserves": 16,
    "treasuries": 16, "repo": 14, "money_mkt_funds": 16,
    "cash": 14, "settlement_adj": 15, "secured_loans": 14,
    "bitcoin": 14, "gold": 12, "non_us_treasuries": 18, "buidl": 14,
    "token": 8, "counterparty": 14, "asset_type": 22,
    "cusip": 12, "maturity_date": 13, "par_value": 16,
    "fair_value": 16, "days_to_maturity": 17,
    "blockchain": 22,
}


def write_sheet(ws, cols, rows, freeze="A2"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(cols, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            if col in ("report_date", "maturity_date", "attestation_date"):
                cell.number_format = DATE_FMT
            elif col == "source_file":
                pass
            elif col == "days_to_maturity":
                cell.number_format = INT_FMT
            elif val is not None and isinstance(val, (int, float)):
                cell.number_format = INT_FMT if val == int(val) else NUM_FMT


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    coins = load_all_coins()
    for token, rows in coins.items():
        if not rows:
            continue
        ws = wb.create_sheet(token)
        # Drop internal _token key, sort by report_date
        clean = sorted(
            [{c: r.get(c) for c in AGGREGATE_COLS} for r in rows],
            key=lambda r: r.get("report_date") or date.min,
        )
        # Drop columns that are entirely None to keep sheets tidy
        present = [c for c in AGGREGATE_COLS
                   if any(r.get(c) is not None for r in clean)]
        write_sheet(ws, present, clean)
        print(f"  {token:<8} {len(clean):>4} rows  ({len(present)} cols)")

    # Securities sheet
    securities = (
        load_usdc_securities()
        + load_busd_securities()
        + load_pyusd_securities()
    )
    if securities:
        securities.sort(key=lambda r: (r.get("token") or "", r.get("report_date") or date.min))
        ws = wb.create_sheet("Securities")
        write_sheet(ws, SECURITIES_COLS, securities)
        print(f"  {'Securities':<8} {len(securities):>4} rows")

    # USDTB BUIDL detail
    buidl_path = PROCESSED / "usdtb" / "usdtb_buidl_detail.csv"
    if buidl_path.exists():
        buidl_rows = []
        for r in read_csv(buidl_path):
            buidl_rows.append({
                "report_date": to_date(r.get("report_date")),
                "source_file": r.get("file"),
                "blockchain":  r.get("blockchain"),
                "fair_value":  to_num(r.get("fair_value")),
            })
        buidl_rows.sort(key=lambda r: r.get("report_date") or date.min)
        ws = wb.create_sheet("USDTB_BUIDL")
        write_sheet(ws, BUIDL_COLS, buidl_rows)
        print(f"  {'USDTB_BUIDL':<8} {len(buidl_rows):>4} rows")

    wb.save(OUT_FILE)
    print(f"\nSaved -> {OUT_FILE}")


if __name__ == "__main__":
    print("Building unified stablecoin reserves workbook...\n")
    build_workbook()
