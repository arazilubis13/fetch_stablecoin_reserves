"""
Extracts reserve data for manually-collected coins from the master xlsx and
writes one CSV per coin to data/processed/{coin}/.

Coins handled: GUSD, USD1, USDP, EURS, AUDD, AUDM
"""

import csv
from pathlib import Path
from datetime import datetime, date
import openpyxl

ROOT = Path(__file__).parent.parent.parent
XLSX = ROOT / "Stablecoin_Reserves_External.xlsx"
PROCESSED = ROOT / "data" / "processed"

SOURCE = "Stablecoin_Reserves_External.xlsx"

RESERVE_COLS = [
    "report_date", "source_file", "auditor",
    "tokens_in_circulation", "total_reserves",
    "treasuries", "repo", "money_mkt_funds",
    "cash", "settlement_adj",
]


def xl_date(val):
    """Convert Excel serial date or datetime to a Python date."""
    if isinstance(val, (datetime, date)):
        return val.date() if hasattr(val, "date") else val
    if isinstance(val, (int, float)):
        # Excel serial: days since 1899-12-30
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
    return None


def row_to_dict(keys, values):
    return dict(zip(keys, values))


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=RESERVE_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  Wrote {len(rows):,} rows -> {path.name}")


def extract_gusd(wb):
    ws = wb["GUSD"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "report_date":          xl_date(row[0]),
            "source_file":          SOURCE,
            "auditor":              None,
            "tokens_in_circulation":None,
            "total_reserves":       row[6],
            "treasuries":           row[5],   # Treasury Bills
            "repo":                 (row[3] or 0) + (row[4] or 0),  # Overnight + Term Reverse Repo
            "money_mkt_funds":      None,
            "cash":                 (row[1] or 0) + (row[2] or 0),  # Cash + Net Adjustment
            "settlement_adj":       None,
        })
    return rows


def extract_usd1(wb):
    ws = wb["USD1"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "report_date":          xl_date(row[0]),
            "source_file":          SOURCE,
            "auditor":              None,
            "tokens_in_circulation":None,
            "total_reserves":       row[5],   # Total_Reserves
            "treasuries":           row[3],   # Treasuries
            "repo":                 None,
            "money_mkt_funds":      row[4],   # Gov_MMFs
            "cash":                 (row[1] or 0) + (row[2] or 0),  # BitGo + Demand Deposits
            "settlement_adj":       None,
        })
    return rows


def extract_usdp(wb):
    ws = wb["USDP"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "report_date":          xl_date(row[0]),
            "source_file":          SOURCE,
            "auditor":              None,
            "tokens_in_circulation":None,
            "total_reserves":       row[5],   # Total_Reserves
            "treasuries":           row[3],   # Treasuries
            "repo":                 row[1],   # Overnight Repos
            "money_mkt_funds":      None,
            "cash":                 row[2],   # Cash
            "settlement_adj":       None,
        })
    return rows


def extract_other(wb):
    """Other sheet has EURC, EURS, AUDM etc. — split by token."""
    ws = wb["Other"]
    by_token = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        token = str(row[1]).strip().upper()
        r = {
            "report_date":          xl_date(row[0]),
            "source_file":          SOURCE,
            "auditor":              None,
            "tokens_in_circulation":None,
            "total_reserves":       row[5],
            "treasuries":           row[2],   # Government Securities
            "repo":                 None,
            "money_mkt_funds":      None,
            "cash":                 row[3],   # Cash at regulated institutions
            "settlement_adj":       row[4],   # Timing differences
        }
        by_token.setdefault(token, []).append(r)
    return by_token


def extract_audd(wb):
    ws = wb["AUDD"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        total = row[6] if row[6] is not None else sum(
            v for v in row[2:6] if isinstance(v, (int, float))
        )
        rows.append({
            "report_date":          xl_date(row[0]),
            "source_file":          SOURCE,
            "auditor":              None,
            "tokens_in_circulation":None,
            "total_reserves":       total,
            "treasuries":           None,
            "repo":                 None,
            "money_mkt_funds":      None,
            "cash":                 total,    # All holdings are cash at various banks
            "settlement_adj":       None,
        })
    return rows


def main():
    print(f"Reading {XLSX.name} ...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    write_csv(PROCESSED / "gusd" / "gusd_reserves.csv",  extract_gusd(wb))
    write_csv(PROCESSED / "usd1" / "usd1_reserves.csv",  extract_usd1(wb))
    write_csv(PROCESSED / "usdp" / "usdp_reserves.csv",  extract_usdp(wb))
    write_csv(PROCESSED / "audd" / "audd_reserves.csv",  extract_audd(wb))

    other = extract_other(wb)
    for token, rows in other.items():
        slug = token.lower()
        write_csv(PROCESSED / slug / f"{slug}_reserves.csv", rows)

    print("Done.")


if __name__ == "__main__":
    main()
